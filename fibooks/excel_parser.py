__author__ = 'Timo Kats'
__name__ = 'excel_parser.py'
__desciption__ = 'parses excel into dataset'

import openpyxl, warnings
warnings.filterwarnings('ignore')

def clean_text(text):
    return text.lower().replace('  ', '')

def load_excel(filename, max_row=None, max_column=None):
    data = {}
    wookbook = openpyxl.load_workbook(filename, data_only=True)
    worksheet = wookbook.active
    if max_row == None:
        max_row = worksheet.max_row
    if max_column == None:
        max_column = worksheet.max_column

    # get the data into structured variable
    for column in worksheet.iter_cols(1, max_column):
        data[column[0].column - 1] = []
        for row in range(0, max_row):
            data[column[0].column - 1].append(column[row].value)
    return data

def get_accounts(data):
    accounts = [] 
    for x in data.keys():
        potential_indent = False 
        for y, row in enumerate(data[x]):  
            if isinstance(data[x][y],str) and (data[x][y] not in accounts):
                if x < list(data.keys())[-1]:
                    potential_indent = True
                accounts.append({'name':data[x][y],'column':x,'row':y})
            elif (potential_indent): # for indents...
                if isinstance(data[x+1][y],str) and (data[x+1][y] not in accounts): 
                    accounts.append({'name':data[x+1][y],'column':x+1,'row':y})
    return accounts   

def create_dataset(data, accounts):
    result = {}
    for account in accounts:
        key = clean_text(data[account['column']][account['row']])
        result[key] = []
        account_active = False
        for index in range(account['column'], len(data.keys())):
            if isinstance(data[index][account['row']], int):
                account_active = True
                result[key].append(data[index][account['row']])
            elif isinstance(data[index][account['row']], float):
                account_active = True
                result[key].append(round(data[index][account['row']], 2))
            if isinstance(data[index][account['row']], str) and account_active:
                break
        if result[key] == []:
            result.pop(key)
    return result

def import_excel(filename, min_row, max_row, min_column, max_column):
    data = load_excel(filename, max_row=max_row, max_column=max_column)
    accounts = get_accounts(data)
    return create_dataset(data, accounts)
    